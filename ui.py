import tkinter as tk

import math
import re  # for working with hex color codes

from button import PrimaryButton
import pandas as pd
import os
import numpy as np
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class menu:
    def __init__(self, data, parquet_file) -> None:

        #   yaml  data
        self.data = data

        self.parquet_file = parquet_file

        self.df = pd.read_parquet(os.path.abspath(parquet_file))
        self.excel_path = f"{os.path.abspath(parquet_file)}.xlsx"
        self.root = tk.Tk()
        self.set_root_properties()
        self.x_origin, self.y_origin = self.root.winfo_pointerxy()

        #def on_focus_out(event):
        #    # Exit the program
        #    print("LOST FOCUS")
        #    self.root.destroy()
        #
        ## Bind the on_focus_out function to the <FocusOut> event of the main window
        #self.root.bind('<FocusOut>', on_focus_out)


        #  default colors
        default_colors = {
            "bg": "#656565",
            "focus": "#D55A00",
            "button": "#000000",
            "text": "#989898",
            "label": "#BABABA",
            "dark_text": "#323232",
        }


        #  monokai-
        #  Background: (46, 46, 46);  #2e2e2e
        #  Comments: (121, 121, 121); #797979
        #  White: (214, 214, 214);    #d6d6d6
        #  Yellow: (229, 181, 103);   #e5b567
        #  Green: (180, 210, 115);    #b4d273
        #  Orange: (232, 125, 62);    #e87d3e
        #  Purple: (158, 134, 200);   #9e86c8
        #  Pink: (176, 82, 121);      #b05279
        #  Blue: (108, 153, 187);     #6c99bb
        monokai_colors = {
            "bg":           "#2e2e2e",
            "focus":        "#e87d3e",
            "button":       "#000000",
            "text":         "#d6d6d6",
            "label":        "#d6d6d6",
            "dark_text":    "#2e2e2e",
            "start_border": "#b4d273",
        }

        #  SOLARIZED	HEX     TERMCOL	  HEX	 
        #  base03	#002b36     brblack	  #1c1c1c
        #  base02	#073642     black	  #262626
        #  base01	#586e75     brgreen	  #585858
        #  base00	#657b83     bryellow  #626262
        #  base0	#839496     brblue	  #808080
        #  base1	#93a1a1     brcyan	  #8a8a8a
        #  base2	#eee8d5     white	  #e4e4e4
        #  base3	#fdf6e3     brwhite	  #ffffd7
        #  yellow	#b58900     yellow	  #af8700
        #  orange	#cb4b16     brred	  #d75f00
        #  red	    #d30102     red	      #af0000
        #  magenta	#d33682     magenta	  #af005f
        #  violet	#6c71c4     brmagenta #5f5faf
        #  blue	    #268bd2     blue	  #0087ff
        #  cyan	    #2aa198     cyan	  #00afaf
        #  green	#859900     green	  #5f8700

        #  these need some work still
        solarized_colors = {
            "bg":        "#1c1c1c",
            "focus":     "#af8700",
            "button":    "#626262",
            "text":      "#8a8a8a",
            "label":     "#ffffd7",
            "dark_text": "#1c1c1c",
        }

        #self.colors = default_colors
        self.colors = monokai_colors
        #self.colors = solarized_colors


        #  setup canvas
        self.canvas = tk.Canvas(self.root, width=900, height=900)
        self.canvas.pack(fill=tk.BOTH, expand=True)

        ###########  establish primary buttons and labels

        #  calculate the angle of each button around the circumference
        self.oval_count = len(self.data["projects"])
        self.step = 360 / self.oval_count

        if self.oval_count < 8:
            self.primary_radius = 60
            self.oval_scale_factor = 2.5
        elif self.oval_count < 10:
            self.primary_radius = 70
            self.oval_scale_factor = 3
        elif self.oval_count < 12:
            self.primary_radius = 150
            self.oval_scale_factor = 4
        else:
            quit("unsupported number of items on primary ring")
        #  this is the base background and primary circle buttons
        self.draw_default_view()

        #  to be populated with primary button objects
        self.buttons = []
        self.arcs = []
        self.secondary_buttons = {}
        self.secondary_buttons_list = []
        for button_number, project in enumerate(self.data["projects"]):
            
            #  if there are no records for this project yet,
            if self.df[self.df['project_name'] == project['name']].empty:
                #  stop time is one second from now
                one_second_later = datetime.now() + timedelta(seconds=1)
                #  generate a first blank row, this makes other logic easier to handle later
                self.df.loc[len(self.df)] = [project['id'], project['name'], pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), pd.to_datetime(one_second_later.strftime('%Y-%m-%d %H:%M:%S')), np.nan, "starter entry - ignore"]


            #  basic math for this button
            angle = (button_number * self.step) + (
                self.step // 2
            )  #  prevents buttons from forming directly at 0, 90, 120, and 180 markers

            #  x,y here represent the center of this button
            x = self.x_origin + self.primary_radius * math.cos(math.radians(angle))
            y = self.y_origin + self.primary_radius * math.sin(math.radians(angle))

            radian_start = angle - (self.step // 2)  #  for future use
            radian_end = angle + (self.step // 2)  #  for future use

            radius = self.primary_radius / self.oval_scale_factor
            
            self.buttons.append(
                PrimaryButton(
                    project,
                    self.canvas,
                    self.x_origin,
                    self.y_origin,
                    radius,
                    angle,
                    self.primary_radius,
                    self.colors,
                    button_number,
                )
            )

            self.secondary_buttons_list.append([])
            #  functions are handled in the parent class
            self.canvas.tag_bind(
                f"button{button_number}",
                "<Enter>",
                lambda event, project=self.buttons[
                    button_number
                ].data, oval=self.buttons[button_number].oval_id: self.button_on_enter(
                    event, project, oval
                ),
            )

            self.canvas.tag_bind(
                f"button{button_number}",
                "<Leave>",
                lambda event, project=self.buttons[
                    button_number
                ].data, oval=self.buttons[button_number].oval_id: self.button_on_leave(
                    event, project, oval
                ),
            )

            self.canvas.tag_bind(
                f"button{button_number}",
                "<Button-1>",
                lambda event, project=self.buttons[
                    button_number
                ].data, oval=self.buttons[
                    button_number
                ].oval_id: self.handle_button_click(
                    event, project, oval
                ),
            )

            num_circles = 3
            circle_radius = radius

            #  This formula establishes the relation between radius of big circle R,
            #    radius of small circle r and number of (touching) small circles N.
            #    We will need to find R to calculate how much space we need on our arc
            #    in order to fit all the buttons
            #  R = r / Sin(Pi/N)
            radius_container = (circle_radius) / math.sin(math.pi / num_circles)

            #  Now turn the radius into the circumfrence
            # circumfrence_container = ((math.pi * radius_container) * 2) * 0.9
            circumfrence_container = (num_circles * circle_radius) * 1.4
            #  Now that we have the circumfrence, we need to project that onto an arc
            #    To do this we need to find the angle of that arc which will produce
            #    an equal size to the circumference we found above which will fit
            #    all the child circles (buttons)
            ###################
            #  The formula:
            #  L = r * θ
            ###################
            #  Hence, the arc length is equal to radius multiplied by the central angle, θ (in radians).
            #  Arc Length = r x θ
            #  Arc Length / r = θ
            #  θ = Arc Length / r

            arc_radius = 80  #  start small, increase until we can fit all buttons

            #  If the angle required to produce the arc of necessary size is bigger than 120
            #    then we can slowly begin to increase the size of the radius
            #    This will guarantee that we will never have an arc greater than 120 degrees
            arc_angle = math.degrees(circumfrence_container / arc_radius)

            #  Calculate the angles for the start and end points of the arc
            if y < self.y_origin:
                radian_start = angle + (arc_angle / 2) % 360
                radian_end = (angle - (arc_angle / 2)) % 360
            else:
                radian_start = angle - (arc_angle / 2) % 360
                radian_end = (angle + (arc_angle / 2)) % 360

            if x < self.x_origin:  # mirror across 180* mark
                mirror_line = 180
            else:  # mirror across 360* mark
                mirror_line = 360
            
            #  mirror the arc vertically, not sure why this needs to be done - can be investigated later
            radian_start = mirror_line - (radian_start - mirror_line)
            radian_end = mirror_line - (radian_end - mirror_line)

            start_angle = radian_start
            end_angle = radian_end

            while arc_angle > 120:
                arc_radius = arc_radius * 1.2
                #  the answer here is in radians, it needs to be in degrees
                arc_angle = start_angle + math.degrees(
                    circumfrence_container / arc_radius
                )  #  θ = Arc Length / r
            total_degrees = end_angle - start_angle
            center_x = x
            center_y = y

            #if center_y < self.y_origin:
            #    center_y = center_y * -1

            coordinates = [
                center_x - arc_radius,
                center_y - arc_radius,
                center_x + arc_radius,
                center_y + arc_radius,
            ]

            #  Draw the arc using the create_arc method
            arc_id = self.canvas.create_arc(
                coordinates,
                start=start_angle,
                extent=end_angle - start_angle,
                # extent=arc_angle,
                tags=(f"arc{button_number}", f"arc_{project['name']}", "arc", "secondary"),
                outline=self.change_color_brightness(self.colors["bg"], 25),
                width=circle_radius * 1.4,
                fill=self.change_color_brightness(self.colors["bg"], 25),
                state=tk.HIDDEN,
            )

            self.arcs.append(arc_id)


            #  functions are handled in the parent class
            self.canvas.tag_bind(
                f"arc_{project['name']}",
                "<Enter>",
                lambda event, project=project, arc=self.arcs[button_number]: self.arc_on_enter(
                    event, project, arc
                ),
            )

            self.canvas.tag_bind(
                f"arc_{project['name']}",
                "<Leave>",
                lambda event, project=project, arc=self.arcs[button_number]: self.arc_on_leave(
                    event, project, arc
                ),
            )


            #  Calculate gap in degrees for each button
            step = total_degrees / (num_circles - 1)
            small_button_labels = ['Start', 'Stop', 'Remark']
            
            for i in range(num_circles):
                # angle = (i * total_degrees / num_circles) + start_angle
                angle = start_angle + (i * step)
                x = center_x + arc_radius * math.cos(math.radians(angle))
                y = center_y - arc_radius * math.sin(math.radians(angle))


                #  Angle from origin, length to extend past parent button
                # x1, y1 = 50, 150  # <--  the origin
                length = 50
                angle_radians = math.radians(angle) * -1    #  why is this a -1?  it is a mystery
                x2 = x + length * math.cos(angle_radians)
                y2 = y + length * math.sin(angle_radians)

                #  second line segment that will carry a text label
                length = len(project['name']) * 10  #  assumes 10 pixels per character
        
                if x < self.x_origin:
                    length = length * -1
                angle_radians = math.radians(0)  #  at angle 0
                x3 = x2 + length * math.cos(angle_radians)
                y3 = y2 + length * math.sin(angle_radians)

                #  multi-segmented line with join style of round and cap style of round
                label_line_id = self.canvas.create_line(
                    x,
                    y,
                    x2,
                    y2,
                    x3,
                    y3,
                    width=circle_radius + 15,
                    state=tk.HIDDEN,
                    fill=self.colors["label"],
                    joinstyle="round",
                    capstyle="round",
                    tags=(f"secondary_line_{project['name']}", "label", "secondary"),
                )

                #  The label originates from the beginning of the second segment
                label_text_id = self.canvas.create_text(
                    (x2 + x3) / 2,
                    y2,
                    fill=self.colors["dark_text"],
                    state=tk.HIDDEN,
                    font="Times 16 bold",
                    text=small_button_labels[i],
                    tags=(f"secondary_line_{project['name']}", "text", "secondary"),
                )

                #  check if last row has start and stop times that are equal, and if this is a 'Start' button
                last_row = self.df[self.df['project_name'] == project['name']].tail(1)
                if small_button_labels[i] == 'Start' and last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
                    #  if true we want this button to look disabled
                    fill = self.change_color_brightness(self.colors["bg"], 25)
                    outline = self.change_color_brightness(self.colors["bg"], 25)
                else:
                    fill = self.colors["button"]
                    outline = self.colors["bg"]

                secondary_button = self.canvas.create_oval(
                    x - circle_radius,
                    y - circle_radius,
                    x + circle_radius,
                    y + circle_radius,
                    tags=(f"button_{project['name']}", "button", "secondary", 
                          f"button_{project['name']}_{small_button_labels[i]}",
                          f"button_{button_number}_{small_button_labels[i]}"
                          #f"button{self.secondary_buttons_list[button_number][i]}",
                          
                          ),

                    state=tk.HIDDEN,
                    outline=outline,
                    width=5,
                    fill=fill,
                )

                self.canvas.addtag_withtag(f"button{secondary_button}", secondary_button)

                self.secondary_buttons_list[button_number].append(secondary_button)

                self.canvas.create_text(
                    x,
                    y,
                    fill=self.colors["text"],
                    state=tk.HIDDEN,
                    font="Times 16 bold",
                    text=small_button_labels[i][0:3],
                    tags=(f"button_text", f"button_{project['name']}_text", "secondary"),
                )

                self.canvas.tag_bind(
                    f"button{self.secondary_buttons_list[button_number][i]}",
                    "<Enter>",
                    lambda event, project=project, button_type=small_button_labels[i], button=self.secondary_buttons_list[button_number][i]: self.secondary_button_on_enter(
                        event, project, button_type, button
                    ),
                )

                self.canvas.tag_bind(
                    f"button{self.secondary_buttons_list[button_number][i]}",
                    "<Leave>",
                    lambda event, project=project, button_type=small_button_labels[i], button=self.secondary_buttons_list[button_number][i]: self.secondary_button_on_leave(
                        event, project, button_type, button
                    ),
                )

                self.canvas.tag_bind(
                    f"button{self.secondary_buttons_list[button_number][i]}",
                    "<Button-1>",
                    lambda event, project=project, button_type=small_button_labels[i], button=self.secondary_buttons_list[button_number][i]: self.handle_secondary_button_click(
                        event, project, button_type, button
                    ),
                )



        # place all buttons over arcs
        self.canvas.tag_raise("text", "label")
        self.canvas.tag_raise("arc", "label")
        self.canvas.tag_raise("button", "arc")
        self.canvas.tag_raise("button_text", "button")
        self.open_ui()

    def draw_default_view(self) -> None:
        """draws the initial menu as it would look on initial open"""

        # self.draw_border()
        self.draw_inner_oval()



    def button_on_enter(self, event, project, oval_id) -> None:
        print(f"ENTERED BUTTON OBJECT ID {oval_id}")

        #  primary items hidden, button gains focus color
        self.canvas.itemconfigure(oval_id, fill=self.colors["focus"])
        self.canvas.itemconfigure("primary_label", state="hidden")

        #  secondary items appear
        self.canvas.itemconfigure(f"arc_{project['name']}", state="normal")
        self.canvas.itemconfigure(f"button_{project['name']}", state="normal")
        self.canvas.itemconfigure(f"button_{project['name']}_text", state="disabled")
        self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="disabled")

        #  check if last row has start and stop times that are equal
        last_row = self.df[self.df['project_name'] == project['name']].tail(1)
        if last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
            #  if true we want this button to look disabled
            self.canvas.itemconfigure(f"button_{project['name']}_Start", state="disabled", fill=self.change_color_brightness(self.colors["bg"], 25), outline=self.change_color_brightness(self.colors["bg"], 25))


    def secondary_button_on_enter(self, event, project, button_type, oval_id) -> None:
        print(f"ENTERED BUTTON OBJECT ID {oval_id}")

        #  primary items hidden, button gains focus color
        self.canvas.itemconfigure(oval_id, fill=self.colors["focus"])
        self.canvas.itemconfigure("secondary_label", state="hidden")

        #  secondary items appear
        self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="hidden")

        #  check if last row has start and stop times that are equal
        last_row = self.df[self.df['project_name'] == project['name']].tail(1)
        if last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
            #  if true we want this button to look disabled
            self.canvas.itemconfigure(f"button_{project['name']}_Start", fill=self.change_color_brightness(self.colors["bg"], 25), outline=self.change_color_brightness(self.colors["bg"], 25))


    def button_on_leave(self, event, project, oval_id) -> None:
        print(f"LEFT OBJECT ID {oval_id}")

        #  regular button color returns, original lables appear
        self.canvas.itemconfigure(oval_id, fill=self.colors["button"])
        self.canvas.itemconfigure("primary_label", state="disabled")

        #  hide secondary objects again
        self.canvas.itemconfigure(f"arc_{project['name']}", state="hidden")
        self.canvas.itemconfigure(f"button_{project['name']}", state="hidden")
        self.canvas.itemconfigure(f"button_{project['name']}_text", state="hidden")
        self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="hidden")

    def secondary_button_on_leave(self, event, project, button_type, oval_id) -> None:
        print(f"LEFT BUTTON OBJECT ID {oval_id}")



        #  primary items hidden, button gains focus color
        self.canvas.itemconfigure(oval_id, fill=self.colors["button"])

#        self.canvas.itemconfigure("secondary_label", state="hidden")
#
#        #  secondary items appear
#        self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="hidden")

        #  check if last row has start and stop times that are equal
        last_row = self.df[self.df['project_name'] == project['name']].tail(1)
        if last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
            #  if true we want this button to look disabled
            self.canvas.itemconfigure(f"button_{project['name']}_Start", fill=self.change_color_brightness(self.colors["bg"], 25), outline=self.change_color_brightness(self.colors["bg"], 25))



    def arc_on_enter(self, event, project, arc_id) -> None:
        print(f"ENTERED ARC OBJECT ID {arc_id}")

        #  all primary labels hide
        self.canvas.itemconfigure(f"arc_{project['name']}", state="normal")
        self.canvas.itemconfigure(f"button_{project['name']}", state="normal")
        self.canvas.itemconfigure(f"button_{project['name']}_text", state="disabled")
        self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="disabled")
        #self.canvas.itemconfigure(f"button_{project['name']}", state="normal")
        #self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="disabled")
        self.canvas.itemconfigure("primary_label", state="hidden")

        #  check if last row has start and stop times that are equal
        last_row = self.df[self.df['project_name'] == project['name']].tail(1)
        if last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
            #  if true we want this button to look disabled
            self.canvas.itemconfigure(f"button_{project['name']}_Start", fill=self.change_color_brightness(self.colors["bg"], 25), outline=self.change_color_brightness(self.colors["bg"], 25))



    def arc_on_leave(self, event, project, arc_id) -> None:
        print(f"LEFT ARC OBJECT ID {arc_id}")

        #  all primary labels return to view
        #self.canvas.itemconfigure(f"button_{project['name']}", state="hidden")
        #self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="hidden")



###   TO DO:    add an enter/leave for primary oval that resets the view, make secondary buttons correcty do some action


    #   now handled on child button object
    def handle_button_click(self, event, project, oval_id) -> None:
        print(f"primary_oval clicked: {oval_id}, {project['name']}")


    def handle_secondary_button_click(self, event, project, button_type, oval_id) -> None:
        print(f"secondary_oval clicked: {oval_id}, {project['name']}  {button_type}")
        
        project_name =  project['name']
        project_id   =  project['id']
        print("project_name: ", project_name)
        #  find the last row for this project (if one exists)

        last_row = self.df[self.df['project_name'] == project_name].tail(1)



        if button_type == "Start":

            # check if there is any row where start_time is not null and stop_time is empty or null
            #if result:
            if last_row['start_time'].iloc[0] == last_row['stop_time'].iloc[0]:
                print('Start and stop times are equal.')
            else:
                print('Start and stop times are not equal.')
                self.df.loc[len(self.df)] = [project_id, project_name, pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), np.nan, ""]
                #self.df.loc[len(self.df)] = [project_name, pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S')), "", "", ""]

                #  populate remark and stop_time empty or null column entries with empty string
                self.df['remark'] = self.df['remark'].fillna('')
                self.df['stop_time'] = self.df['stop_time'].fillna('')

        elif button_type == "Stop":
            ''' this will update the stop time of a project by selecting the last row of a given project_name if it exists
                it will continue to update the final timestamp of that final row if the button is pressed multiple times
                if the 'Start' button is pressed then a new row is generated
            '''
            #  update the stop_time value for the row - this applies to bottom row that might already have a stop time on it
            self.df.loc[(self.df['project_name'] == project_name)].iloc[-1] = pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            # select the last row for the given project name
            last_row_index = self.df.loc[self.df['project_name'] == project_name].index[-1]
            
            # update the stop time for the last row
            self.df.loc[last_row_index, 'stop_time'] = pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            time_delta_hours = round((self.df.loc[last_row_index, 'stop_time'] - self.df.loc[last_row_index, 'start_time']).seconds / 3600, 2)
            self.df.loc[last_row_index, 'total_time'] = time_delta_hours

            #  let start button be clickable again
            self.canvas.itemconfigure(f"button_{project['name']}_Start", state="normal", fill=self.colors["button"], outline=self.colors["bg"])


#            # check if there is any row where start_time is not null and stop_time is empty or null
#            if result:
#                print("The project has a start time but no stop time")
#
#                #  update the stop_time value for the row - this applies to bottom row that might already have a stop time on it
#                self.df.loc[(self.df['project_name'] == project_name)].iloc[-1] = pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
#                
#                # select the last row for the given project name
#                last_row_index = self.df.loc[self.df['project_name'] == project_name].index[-1]
#                
#                # update the stop time for the last row
#                self.df.loc[last_row_index, 'stop_time'] = pd.to_datetime(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
#
#            else:
#                print("The project has no start time or has both start and stop times")


        elif button_type == "Remark":
           # select the last row for the given project name
            last_row_index = self.df.loc[self.df['project_name'] == project_name].index[-1]
            
            remark_text = self.df.loc[last_row_index, 'remark']

            # Get the x and y coordinates of the mouse click
            x = event.x
            y = event.y
            
            # Create a non-modal dialog window
            dialog = tk.Toplevel(self.root)
            dialog_width = 400
            dialog_height = 100
            dialog.geometry("{}x{}+{}+{}".format(dialog_width, dialog_height, x, y))
            # Add a label and text input field to the dialog window
            #label = tk.Label(dialog, text=)
            label = tk.Label(dialog, text="Add or change the comment for this entry", font=("Helvetica", 16))#, justify=tk.LEFT)
            label.pack()
    
            
            var = tk.StringVar(value=remark_text)
    
            # Remove the title bar of the dialog window
            dialog.overrideredirect(True)
        
            text = tk.Text(dialog, width=50, height=5)
            text.insert('1.0', remark_text) # Set default value
            text.pack(pady=10)
    
            #entry = tk.Entry(dialog, width=50, height=20, textvariable=tk.StringVar(value='DEFAULT'))
            #entry.pack(pady=10)
            
            def handle_ok():
                # Get the value entered by the user
                value = text.get('1.0', 'end-1c')
                
                # Print the value entered by the user
                print(value.strip())
                self.df.loc[last_row_index, 'remark'] = value.strip().replace("\n", "")
                # Close the dialog window
                dialog.destroy()
                
            # Add an OK button to the dialog window
            #ok_button = tk.Button(dialog, text="OK", command=handle_ok, width=10, height=2)
            #ok_button.pack(side=tk.LEFT, padx=10, pady=10)
    
            def handle_cancel():
                # Close the dialog window
                dialog.destroy()
    
            #cancel_button = tk.Button(dialog, text="Cancel", command=handle_cancel, width=10, height=2)
            #cancel_button.pack(side=tk.RIGHT, padx=10, pady=10)
            dialog.bind('<Return>', lambda event: handle_ok())
            dialog.bind('<Escape>', lambda event: handle_cancel())
    
            dialog.geometry("+{}+{}".format(x, y))


        print(self.df)
        self.df.to_parquet(self.parquet_file, index=False)

        # Read Parquet file into a DataFrame
        temp_df = pd.read_parquet(self.parquet_file)
        
        # Write DataFrame to an Excel file
        temp_df.to_excel('file.xlsx', index=False)

        ##  export the DataFrame to an Excel file with each month in its own sheet
        #writer = pd.ExcelWriter('data.xlsx', engine='xlsxwriter')
        #for month in range(1, 13):
        #    sheet_name = datetime.date(1900, month, 1).strftime('%B')
        #    df_month = self.df[self.df['start_time'].dt.month == month]
        #    df_month.to_excel(writer, sheet_name=sheet_name, index=False)
        #writer.save()




    def root_enter(self, print_item):
        print("entered root")

    def root_leave(self, print_item):
        print("left root")
        #  regular button color returns, original lables appear
        #self.canvas.itemconfigure(oval_id, fill=self.colors["button"])
        self.canvas.itemconfigure("primary", state="disabled")
        self.canvas.itemconfigure("primary_buttons", state="normal")

        #  hide secondary objects again
        self.canvas.itemconfigure(f"secondary", state="hidden")
        #self.canvas.itemconfigure(f"button_{project['name']}", state="hidden")
        #self.canvas.itemconfigure(f"secondary_line_{project['name']}", state="hidden")

    def open_ui(self) -> None:
        """draw main ui component"""
        self.root.mainloop()  # <---  must happen last in draw sequence

    def draw_inner_oval(self) -> None:
        """draws just the inner oval, everything else lays on top of this"""

        self.main_oval_id = self.canvas.create_oval(
            self.x_origin - self.primary_radius,  #  top left x position
            self.y_origin - self.primary_radius,  #  top left y position
            self.x_origin + self.primary_radius,  #  bottom right x position
            self.y_origin + self.primary_radius,  #  bottom right y position
            outline=self.colors["bg"],
            width=15,
            # outline="",
            fill=self.colors["bg"],
        )


    def set_root_properties(self) -> None:
        """setup root object attributes. The root object has no concept of x,y coordinates - just make it full screen and transparent"""

        # Hide the UI initially
        ############self.root.withdraw()
        self.root.overrideredirect(True)

        #  make root transparent when showing the 'bg' color
        self.root.wm_attributes("-transparentcolor", self.root["bg"])

        #  root is full screen, but unnoticeable because it is transparent
        self.root.geometry(
            f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}"
        )

        #  escape key quits the UI after writing excel file to disk
        self.root.bind("<Escape>", self.make_excel)

        #self.root.bind("<KeyPress-e>", self.make_excel)


        self.root.bind ("<Enter>", self.root_enter)
        self.root.bind("<Leave>", self.root_leave)


        #  Removes title bar, icon from system tray
        self.root.deiconify()

        #  update all root properties
        self.root.update_idletasks()


    def make_excel(self, event) -> None:

        # Create a new Excel workbook
        wb = Workbook()
        
        # Loop through each unique month in the stop_time column and create a new sheet for each one
        for month in self.df["stop_time"].dt.month.unique():
            
            #  filter the dataframe to only contain rows for the current month
            df_month = self.df[self.df["stop_time"].dt.month == month]
            
            #  create a new sheet with the name of the current month
            sheet_name = pd.Timestamp(year=df_month["stop_time"].iloc[0].year, month=month, day=1).strftime("%B %Y")
            ws = wb.create_sheet(sheet_name)
            
            #  write the column headers to the sheet
            for col_idx, col_name in enumerate(df_month.columns, 1):
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}1"] = col_name
            
            #  write the data to the sheet
            for row_idx, row_data in enumerate(df_month.values, 2):
                for col_idx, cell_data in enumerate(row_data, 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{row_idx}"] = cell_data
        
        #  delete this extra sheet that is generated
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        #  save the Excel workbook
        wb.save(self.excel_path)
        print("excel file saved")
        self.root.quit()

    def change_color_brightness(
        self, hex_color_code: str, brightness_change: int
    ) -> str:
        """takes a current hex color code and an integer used to adjust brightness.
        a negative brightness change makes the color darker
        a positive brightness change makes the color lighter"""

        #  Extract only hexadecimal characters from the input string
        hex_color_code = re.sub("[^0-9a-fA-F]", "", hex_color_code)

        #  Convert the hex color code to RGB values
        red = int(hex_color_code[0:2], 16)
        green = int(hex_color_code[2:4], 16)
        blue = int(hex_color_code[4:6], 16)

        #  Calculate the new RGB values based on the brightness change - max of 255 and min of 0
        new_red = max(min(red + brightness_change, 255), 0)
        new_green = max(min(green + brightness_change, 255), 0)
        new_blue = max(min(blue + brightness_change, 255), 0)

        #  Convert the new RGB values back to a hex color code and return it
        return "#{:02x}{:02x}{:02x}".format(new_red, new_green, new_blue)
